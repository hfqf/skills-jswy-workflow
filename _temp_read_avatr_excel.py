import openpyxl

workbook_path = r"c:\git\skills-jswy-workflow\templates\阿维塔\江苏万友上海阿维塔经营简化执行方案.xlsx"
wb = openpyxl.load_workbook(workbook_path, data_only=False)

print("SHEETS:", wb.sheetnames)

top_rows = 32
top_cols = 22

sample_formula_rows = {
    "01_3月经营总表": [8, 12, 13, 14, 15, 16, 19, 20, 21, 22, 23],
    "02_线索来源分解": [6, 7, 22, 26],
    "03_门店顾问分解": [6, 27],
    "04_周执行追踪": [6],
    "05_销售分析排名": [5, 6, 7, 8, 12, 35, 41],
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print("\nSHEET", sheet_name, "TOP_ROWS")
    for r in range(1, top_rows + 1):
        row_values = []
        has_value = False
        for c in range(1, top_cols + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                has_value = True
            row_values.append(val)
        if has_value:
            print(str(r) + ": " + str(row_values))

    print("\nSHEET", sheet_name, "SAMPLE_FORMULAS")
    rows = sample_formula_rows.get(sheet_name, [])
    for r in rows:
        for c in range(1, top_cols + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                print(cell.coordinate + ": " + val)
